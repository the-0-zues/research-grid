"""Export DER bus placements per scenario to Excel."""
import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIG = r"C:\EPRI_Ckt7_DER_Project\simulation_config.json"
OUT    = r"C:\EPRI_Ckt7_DER_Project\Simulation_Results\DER_Bus_Placements.xlsx"

with open(CONFIG) as f:
    config = json.load(f)

scenarios = config["scenarios"]

wb = openpyxl.Workbook()

# ── Sheet 1: Summary table (scenario x new buses) ──
ws1 = wb.active
ws1.title = "By Scenario"

# Styles
hdr_fill  = PatternFill("solid", fgColor="1F4E79")
hdr_font  = Font(bold=True, color="FFFFFF", size=11)
new_fill  = PatternFill("solid", fgColor="E2EFDA")   # light green = new buses
cum_fill  = PatternFill("solid", fgColor="DDEEFF")   # light blue  = carried over
alt_fill  = PatternFill("solid", fgColor="F5F5F5")
center    = Alignment(horizontal="center", vertical="center")
thin      = Side(style="thin", color="CCCCCC")
border    = Border(left=thin, right=thin, top=thin, bottom=thin)

max_buses = max(len(s["all_buses"]) for s in scenarios)

# Header row
headers = ["Scenario", "Label", "Cumulative DG Count", "New Buses This Step"] + \
          [f"DG Bus #{i+1}" for i in range(max_buses)]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font      = hdr_font
    cell.fill      = hdr_fill
    cell.alignment = center
    cell.border    = border

# Data rows
for r, sc in enumerate(scenarios, 2):
    fill = alt_fill if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    new_set = set(sc["new_buses"])

    ws1.cell(row=r, column=1, value=sc["scenario_id"]).border  = border
    ws1.cell(row=r, column=2, value=sc["label"]).border        = border
    ws1.cell(row=r, column=3, value=len(sc["all_buses"])).border = border
    ws1.cell(row=r, column=4, value=", ".join(sc["new_buses"])).border = border

    for col, bus in enumerate(sc["all_buses"], 5):
        cell        = ws1.cell(row=r, column=col, value=bus)
        cell.border = border
        cell.alignment = center
        cell.fill   = new_fill if bus in new_set else fill

    # Style fixed columns
    for col in range(1, 5):
        ws1.cell(row=r, column=col).fill      = fill
        ws1.cell(row=r, column=col).alignment = center

# Column widths
ws1.column_dimensions["A"].width = 10
ws1.column_dimensions["B"].width = 14
ws1.column_dimensions["C"].width = 22
ws1.column_dimensions["D"].width = 40
for i in range(5, 5 + max_buses):
    ws1.column_dimensions[get_column_letter(i)].width = 12

ws1.freeze_panes = "A2"
ws1.row_dimensions[1].height = 20

# ── Sheet 2: Flat list (bus, scenario first introduced, all scenarios present) ──
ws2 = wb.create_sheet("Bus Introduction Order")

hdr2 = ["DG Bus", "Introduced At Scenario", "Introduced At Label",
        "Introduced At DG Count", "Present In Scenarios (cumulative from)"]
for col, h in enumerate(hdr2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font      = hdr_font
    cell.fill      = hdr_fill
    cell.alignment = center
    cell.border    = border

# Build bus -> first scenario mapping
bus_intro = {}
for sc in scenarios:
    for bus in sc["new_buses"]:
        if bus not in bus_intro:
            bus_intro[bus] = sc

for r, (bus, sc) in enumerate(bus_intro.items(), 2):
    fill = alt_fill if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    vals = [bus, sc["scenario_id"], sc["label"], len(sc["all_buses"]),
            f"Scenario {sc['scenario_id']} ({sc['label']}) onwards"]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=col, value=v)
        cell.fill      = fill
        cell.border    = border
        cell.alignment = center

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 24
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 24
ws2.column_dimensions["E"].width = 36
ws2.freeze_panes = "A2"

wb.save(OUT)
print(f"Saved -> {OUT}")
print(f"  Sheet 1: {len(scenarios)} scenarios x up to {max_buses} DG buses")
print(f"  Sheet 2: {len(bus_intro)} unique DG buses with introduction order")
print("  Green cells = newly added buses at that step")
print("  Blue cells  = carried over from previous scenario")
